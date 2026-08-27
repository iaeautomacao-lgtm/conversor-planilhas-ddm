import csv
from openpyxl import load_workbook
from .text import normalize_header, find_column_index

def detectar_encoding(path: str) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            with open(path, "r", encoding=enc) as f:
                f.read(4096)
            return enc
        except Exception:
            continue
    return "cp1252"

def detectar_delimitador(sample: str) -> str:
    count_semicolon = sample.count(";")
    count_comma = sample.count(",")
    count_tab = sample.count("\t")
    if count_tab > count_semicolon and count_tab > count_comma:
        return "\t"
    return ";" if count_semicolon >= count_comma else ","

def read_input_rows(file_path: str, ext: str):
    ext = ext.lower()
    if ext == ".xlsx":
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], []
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        headers_norm = [normalize_header(h) for h in headers]
        
        raw_rows = []
        for r in rows_iter:
            row_dict = {}
            for idx, h in enumerate(headers):
                if h:
                    val = r[idx] if idx < len(r) else ""
                    row_dict[h] = val if val is not None else ""
            raw_rows.append(row_dict)
        return headers, raw_rows
    else:
        enc = detectar_encoding(file_path)
        with open(file_path, "r", encoding=enc, newline="") as f:
            sample = f.read(4096)
            delimiter = detectar_delimitador(sample)
            f.seek(0)
            reader = csv.reader(f, delimiter=delimiter)
            try:
                header_row = next(reader)
            except StopIteration:
                return [], []
            headers = [str(h).strip() if h is not None else "" for h in header_row]
            raw_rows = []
            for r in reader:
                row_dict = {}
                for idx, h in enumerate(headers):
                    if h:
                        val = r[idx] if idx < len(r) else ""
                        row_dict[h] = val
                raw_rows.append(row_dict)
            return headers, raw_rows
